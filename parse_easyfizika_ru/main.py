import requests
import bs4
import json
from fake_headers import Headers
import re
from pprint import pprint
from openpyxl import Workbook
from openpyxl import load_workbook




def get_headers():
    return Headers(os='win', browser='chrome').generate()


def download_data_to_excel(data):
    #wb = Workbook()
    wb = load_workbook('tasks_easyfizika.xlsx')
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save('tasks_easyfizika.xlsx')

# wb = load_workbook('tasks_easyfizika.xlsx')
# ws = wb.active
# ws.insert_rows(idx=1)
# ws['A1'] = 'Номер'
# ws['B1'] = 'Условие'
# ws['C1'] = 'Рисунки к условию'
# ws['D1'] = 'Решение'
# ws['E1'] = 'Рисунки к решению'
# ws['F1'] = 'Тема'
# ws['G1'] = 'Подтема'
# wb.save('tasks_easyfizika.xlsx')

BASE_URL = 'https://easyfizika.ru/zadachi/kvanty-atom-atomnoe-yadro/'
parsed_data = []
# wb = Workbook()
# ws = wb.active
# data = [('Номер', 'Условие', 'Рисунки к условию', 'Решение', 'Рисунки к решению', 'Тема', 'Подтема')]
# for row in data:
#     ws.append(row)
# wb.save('tasks_easyfizika.xlsx')
response = requests.get(BASE_URL, headers = get_headers())
main_html_data = response.text
main_soup = bs4.BeautifulSoup(main_html_data, features='lxml')
main_page = main_soup.find('div', class_='entry-content')
subtopics = main_page.find_all('h3')
tasks_blocks = main_page.find_all('p')
pat_numb = re.compile(r' \d{1,2}\.\d{1,2}\.\d{1,3}')
for id, elem in enumerate(tasks_blocks):
    a_tags = elem.find_all('a')
    task_images = []
    data_text = []
    tasks_ans = []
    for idx, a_tag in enumerate(a_tags):
        print(id, idx)
        task_href = a_tag['href']
        response = requests.get(BASE_URL + task_href, headers=get_headers())
        task_html_data = response.text
        task_soup = bs4.BeautifulSoup(task_html_data, features='lxml')
        task_deep = task_soup.find('div', class_='entry-content')
        task_text_blocks = task_deep.find_all('p')
        task_img_cond = ''
        task_img_sol = ''
        for t in task_text_blocks:
            if t.find('a'):
                if 'https://easyfizika.ru/wp-content/uploads' in t.find('a')['href']:
                    if 'resheniyu' in t.find('a')['href']:
                        if task_img_sol == '':
                            task_img_sol += t.find('a')['href']
                        else:
                            task_img_sol += '\n'
                            task_img_sol += t.find('a')['href']
                    if 'usloviyu' in t.find('a')['href']:
                        if task_img_cond == '':
                            task_img_cond += t.find('a')['href']
                        else:
                            task_img_cond += '\n'
                            task_img_cond += t.find('a')['href']
        task_images.append((task_img_cond, task_img_sol))
        pat_tags = re.compile(r'<.+>$')
        good_text = []
        flag_dano_exists = 0
        for it in task_deep.prettify().split('\n'):
            if not pat_tags.findall(it):
                good_text.append(it)
            if 'Дано:' in it:
                flag_dano_exists = 1
        task_text = ''
        task_solution = ''
        task_ans = ''
        flag_ans = 0
        flag_dano = 0
        flag_sol = 0
        flag_final = 0
        pat_degree = re.compile(r'^   \d$')
        for idr, row in enumerate(good_text):
            if 'Дано:' in row:
                flag_dano = 1
            if 'Решение задачи:' in row:
                flag_sol = 1
            if 'Если Вы не поняли решение' in row:
                flag_final = 1
            if flag_dano_exists == 1:
                if 'Задача №' not in row and 'Условие задачи:' not in row and flag_dano == 0:
                    if len(row) > 0:
                        if pat_degree.findall(row):
                            task_text += '\\(^'
                            task_text += row[3:]
                            task_text += '\\)'
                        else:
                            if row[0] == ' ' and row[1] == ' ':
                                task_text += row[2:]
                                continue
                            if row[0] == ' ' and row[1] != ' ':
                                task_text += row[1:]
                                continue
                            if row[0] != ' ':
                                task_text += row
                                continue
                if 'Ответ:' in row:
                    flag_ans = 1
                    if row[0] == ' ' and row[1] == ' ':
                        task_ans += row[2:]
                    if row[0] == ' ' and row[1] != ' ':
                        task_ans += row[1:]
                    if row[0] != ' ':
                        task_ans += row
                if flag_dano == 1 and flag_sol == 0 and flag_ans == 0:
                    if pat_degree.findall(row):
                        task_solution += '\\(^'
                        task_solution += row[3:]
                        task_solution += '\\)'
                    else:
                        if row[0] == ' ' and row[1] == ' ':
                            task_solution += row[2:]
                        if row[0] == ' ' and row[1] != ' ':
                            task_solution += row[1:]
                        if row[0] != ' ':
                            task_solution += row
                if flag_dano == 1 and flag_sol == 1 and flag_ans == 0:
                    if pat_degree.findall(row):
                        task_solution += '\\(^'
                        task_solution += row[3:]
                        task_solution += '\\)'
                    else:
                        if task_solution == '':
                            if row[0] == ' ' and row[1] == ' ':
                                task_solution += row[2:]
                            if row[0] == ' ' and row[1] != ' ':
                                task_solution += row[1:]
                            if row[0] != ' ':
                                task_solution += row
                        else:
                            task_solution += '\n'
                            if row[0] == ' ' and row[1] == ' ':
                                task_solution += row[2:]
                            if row[0] == ' ' and row[1] != ' ':
                                task_solution += row[1:]
                            if row[0] != ' ':
                                task_solution += row
                if flag_ans == 1 and flag_final == 0 and 'Ответ:' not in row:
                    if(len(row) > 0):
                        if pat_degree.findall(row):
                            task_ans += '\\(^'
                            task_ans += row[3:]
                            task_ans += '\\)'
                        else:
                            if row[0] == ' ' and row[1] == ' ':
                                task_ans += row[2:]
                            if row[0] == ' ' and row[1] != ' ':
                                task_ans += row[1:]
                            if row[0] != ' ':
                                task_ans += row
            if flag_dano_exists == 0:
                if 'Задача №' not in row and 'Условие задачи:' not in row and flag_sol == 0:
                    if len(row) > 0:
                        if pat_degree.findall(row):
                            task_text += '\\(^'
                            task_text += row[3:]
                            task_text += '\\)'
                        else:
                            if row[0] == ' ' and row[1] == ' ':
                                task_text += row[2:]
                                continue
                            if row[0] == ' ' and row[1] != ' ':
                                task_text += row[1:]
                                continue
                            if row[0] != ' ':
                                task_text += row
                                continue
                if 'Ответ:' in row:
                    flag_ans = 1
                    if row[0] == ' ' and row[1] == ' ':
                        task_ans += row[2:]
                    if row[0] == ' ' and row[1] != ' ':
                        task_ans += row[1:]
                    if row[0] != ' ':
                        task_ans += row
                if flag_sol == 1 and flag_ans == 0:
                    if pat_degree.findall(row):
                        task_solution += '\\(^'
                        task_solution += row[3:]
                        task_solution += '\\)'
                    else:
                        if task_solution == '':
                            if row[0] == ' ' and row[1] == ' ':
                                task_solution += row[2:]
                            if row[0] == ' ' and row[1] != ' ':
                                task_solution += row[1:]
                            if row[0] != ' ':
                                task_solution += row
                        else:
                            task_solution += '\n'
                            if row[0] == ' ' and row[1] == ' ':
                                task_solution += row[2:]
                            if row[0] == ' ' and row[1] != ' ':
                                task_solution += row[1:]
                            if row[0] != ' ':
                                task_solution += row
                if flag_ans == 1 and flag_final == 0 and 'Ответ:' not in row:
                    if(len(row) > 0):
                        if pat_degree.findall(row):
                            task_ans += '\\(^'
                            task_ans += row[3:]
                            task_ans += '\\)'
                        else:
                            if row[0] == ' ' and row[1] == ' ':
                                task_ans += row[2:]
                            if row[0] == ' ' and row[1] != ' ':
                                task_ans += row[1:]
                            if row[0] != ' ':
                                task_ans += row
        task_sol_and_ans = task_solution + '\n' + task_ans
        data_text.append((task_text, task_sol_and_ans))
        tasks_ans.append(task_ans)
    numbers = []
    for item in elem.prettify().split('\n'):
        if pat_numb.findall(item):
            numbers.append(item[1:])
    for i, elem in enumerate(data_text):
        parsed_data.append((numbers[i], elem[0], task_images[i][0],
                            elem[1], task_images[i][1], 'Кванты, атом, атомное ядро', subtopics[id].text))
download_data_to_excel(parsed_data)
