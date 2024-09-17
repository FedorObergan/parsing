import requests
import bs4
import json
from fake_headers import Headers
import re
from pprint import pprint
from openpyxl import Workbook
from openpyxl import load_workbook




def get_headers():
    return Headers(os='win', browser='chrome').generate()


def download_data_to_excel(data):
    #wb = Workbook()
    wb = load_workbook('earthz_tasks_physics.xlsx')
    #wb = load_workbook('physics_themes.xlsx')
    ws = wb.active
    for row in data:
        ws.append(row)
    #wb.save('physics_themes.xlsx')
    wb.save('earthz_tasks_physics.xlsx')


# wb = load_workbook('earthz_tasks_physics.xlsx')
# ws = wb.active
# tasks_list = tuple(ws.iter_rows(values_only=True))
# tasks_list_evolved = tasks_list[1:]
#
#
# wb = load_workbook('physics_themes.xlsx')
# ws = wb.active
# number_theme_list = tuple(ws.iter_rows(values_only=True))
#
# wb = load_workbook('earthz_tasks_physics.xlsx')
# ws = wb.active
# ws['F1'] = 'Тема'
# for id, elem in enumerate(tasks_list_evolved):
#     for numb in number_theme_list:
#         if numb[0] == elem[0]:
#             ws[f'F{id+2}'] = numb[1]
# wb.save('earthz_tasks_physics.xlsx')


# Парсинг тем задач
# BASE_URL = 'https://earthz.ru'
# parsed_data = []
# index = 0
# for i in range(1,41):
#     Url = f'https://earthz.ru/solves~{i}~t13'
#     response = requests.get(Url, headers = get_headers())
#     main_html_data = response.text
#     main_soup = bs4.BeautifulSoup(main_html_data, features='lxml')
#     tasks_tags = main_soup.find_all('div', class_='post-right-solve')
#     for task in tasks_tags:
#         task_href = task.find('a')['href']
#         response = requests.get(BASE_URL+task_href, headers=get_headers())
#         task_html_data = response.text
#         task_soup = bs4.BeautifulSoup(task_html_data, features='lxml')
#         task_numb = task_soup.find('div', attrs={'id': 'right'})
#         task_number = task_numb.find('h2')
#         pat_numb = re.compile(r'\d+')
#         try:
#             a = task_number.prettify().split('\n')
#         except AttributeError:
#             continue
#         for id, elem in enumerate(task_number.prettify().split('\n')):
#             if ((elem == ' Задача по физике - \r') and (len(pat_numb.findall(task_number.prettify().split('\n')[id+1]))>0)):
#                 elem += task_number.prettify().split('\n')[id+1]
#             if 'Задача по' in elem:
#                 number = pat_numb.findall(elem)
#                 index = int(number[0])
#         parsed_data.append((index, 'Атомная и ядерная физика'))
#
# download_data_to_excel(parsed_data)

# wb = load_workbook('earthz_tasks_physics.xlsx')
# ws = wb.active
# ws['A1'] = 'Номер'
# ws['B1'] = 'Условие'
# ws['C1'] = 'Рисунки к условию'
# ws['D1'] = 'Решение'
# ws['E1'] = 'Рисунки к решению'
# wb.save('earthz_tasks_physics.xlsx')

BASE_URL = 'https://earthz.ru'
parsed_data = []
index = 0
for i in range(1000,1095):
    if i == 1:
        data = [('Номер', 'Условие', 'Рисунки к условию', 'Решение', 'Рисунки к решению')]
        download_data_to_excel(data)
    Url = f'https://earthz.ru/solves~{i}~t1'
    response = requests.get(Url, headers = get_headers())
    main_html_data = response.text
    main_soup = bs4.BeautifulSoup(main_html_data, features='lxml')
    tasks_tags = main_soup.find_all('div', class_='post-right-solve')
    for task in tasks_tags:
        task_href = task.find('a')['href']
        response = requests.get(BASE_URL+task_href, headers=get_headers())
        task_html_data = response.text
        task_soup = bs4.BeautifulSoup(task_html_data, features='lxml')
        task_numb = task_soup.find('div', attrs={'id': 'right'})
        task_number = task_numb.find('h2')
        pat_numb = re.compile(r'\d+')
        try:
            a = task_number.prettify().split('\n')
        except AttributeError:
            continue
        for id, elem in enumerate(task_number.prettify().split('\n')):
            if ((elem == ' Задача по физике - \r') and (len(pat_numb.findall(task_number.prettify().split('\n')[id+1]))>0)):
                elem += task_number.prettify().split('\n')[id+1]
            if 'Задача по' in elem:
                number = pat_numb.findall(elem)
                index = int(number[0])
        task_deep = task_soup.find('div', attrs={'id': 'post_solves'})
        res = []
        image_task_url = ''
        image_solution_url = ''
        flag_solution = 0
        for elem in task_deep.prettify().split('\n'):
            if len(elem)>0:
                if elem[0] == ' ':
                    elem = elem[1:]
                if 'Решение' in elem:
                    flag_solution = 1
                if 'img src=' in elem:
                    if flag_solution == 0:
                        if image_task_url == '':
                            image_task_url += BASE_URL
                            image_task_url += elem.split('\"')[1]
                        else:
                            image_task_url += '\n'
                            image_task_url += BASE_URL
                            image_task_url += elem.split('\"')[1]
                    else:
                        if image_solution_url == '':
                            image_solution_url += BASE_URL
                            image_solution_url += elem.split('\"')[1]
                        else:
                            image_solution_url += '\n'
                            image_solution_url += BASE_URL
                            image_solution_url += elem.split('\"')[1]
                if len(elem)>1:
                    if not ((elem[0] == '<' or elem[1] == '<') and elem[-1] == '>'):
                        res.append(elem)
        task_text = ''
        task_solution = ''
        flag = 0
        for elem in res[1:]:
            if 'Решение:' in elem:
                flag = 1
            if flag == 0:
                task_text += elem
            if flag == 1 and 'Решение:' not in elem:
                task_solution += elem
        if task_text[0] == ' ':
            task_text = task_text[1:]
        if ((len(task_solution) > 0) and (task_solution[0] == ' ')):
            task_solution = task_solution[1:]

        parsed_data.append((index, task_text, image_task_url, task_solution, image_solution_url))

download_data_to_excel(parsed_data)
